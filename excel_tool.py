#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 表格换行单元格拆行工具。

读取 .xlsx/.xlsm 工作簿，按首行表头定位 Commit信息 列；如果该列
字符串单元格包含换行符，则将这一行按换行段展开为多行。

拆分规则：
  - 支持 \\n、\\r\\n、\\r
  - 仅 Commit信息 列参与拆分
  - Commit信息 中的空白行会被忽略，不生成输出行
  - 其他列不参与拆分，生成的每一行都复制原行对应列的值
  - 找不到 Commit信息 表头的工作表会跳过并输出告警

导出补丁列表：
  python3 excel_tool.py export-commits INPUT OUTPUT

  从 Commit信息 列导出 patch_tool.py 可读取的文本列表，每行一个 commit
  条目，默认按首次出现顺序去重。
"""

import argparse
import logging
import sys
from copy import copy
from pathlib import Path
from time import monotonic
from typing import Any, Dict, List, Optional, Sequence, Set


SUPPORTED_SUFFIXES = (".xlsx", ".xlsm")
DEFAULT_PROGRESS_INTERVAL = 10000
COMMIT_INFO_HEADER = "Commit信息"

LOG = logging.getLogger(__name__)


def _require_openpyxl() -> Any:
    """延迟导入 openpyxl，避免 --help 依赖运行时包。"""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError(
            "缺少依赖 openpyxl，请先运行: python3 -m pip install -r requirements.txt"
        )
    return openpyxl


def _normalize_newlines(text: str) -> str:
    """将不同换行符归一为 \\n。"""
    return text.replace("\r\n", "\n").replace("\r", "\n")


def _split_cell_value(value: Any) -> List[Any]:
    """按换行符拆分字符串单元格值，并忽略空白段。"""
    if value is None:
        return []
    if not isinstance(value, str):
        return [value]

    normalized = _normalize_newlines(value)
    return [part for part in normalized.split("\n") if part.strip()]


def _header_text(value: Any) -> str:
    """返回用于表头匹配的文本。"""
    if value is None:
        return ""
    return str(value).strip()


def _find_commit_info_column(
    worksheet: Any, max_column: int, header_name: str = COMMIT_INFO_HEADER
) -> Optional[int]:
    """按首行表头查找 Commit信息 列号。"""
    matched_columns: List[int] = []
    for col_idx in range(1, max_column + 1):
        if _header_text(worksheet.cell(row=1, column=col_idx).value) == header_name:
            matched_columns.append(col_idx)

    if not matched_columns:
        return None
    if len(matched_columns) > 1:
        LOG.warning(
            "工作表存在多个 %s 表头: sheet=%s, columns=%s，使用第一个",
            header_name,
            worksheet.title,
            ",".join(str(col_idx) for col_idx in matched_columns),
        )
    return matched_columns[0]


def _copy_cell_format(source: Any, target: Any) -> None:
    """复制常见单元格格式。"""
    if source is target:
        return
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.hyperlink = copy(source.hyperlink) if source.hyperlink else None
    target.comment = copy(source.comment) if source.comment else None


def _copy_row_format(
    worksheet: Any, source_row: int, target_row: int, max_column: int
) -> None:
    """复制行高、隐藏状态和单元格格式。"""
    _copy_row_dimension(worksheet, source_row, target_row)

    for col_idx in range(1, max_column + 1):
        _copy_cell_format(
            worksheet.cell(row=source_row, column=col_idx),
            worksheet.cell(row=target_row, column=col_idx),
        )


def _copy_row_dimension(worksheet: Any, source_row: int, target_row: int) -> None:
    """复制行高和隐藏状态等行级格式。"""
    if source_row == target_row:
        return

    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed


def _copy_row_values_and_format(
    worksheet: Any,
    source_row: int,
    target_row: int,
    max_column: int,
    commit_column: Optional[int] = None,
    commit_parts: Optional[Sequence[Any]] = None,
    part_idx: int = 0,
) -> None:
    """复制一行格式并写入目标行值。"""
    _copy_row_dimension(worksheet, source_row, target_row)

    for col_idx in range(1, max_column + 1):
        source = worksheet.cell(row=source_row, column=col_idx)
        target = worksheet.cell(row=target_row, column=col_idx)
        _copy_cell_format(source, target)
        if commit_parts is not None and col_idx == commit_column:
            target.value = _value_for_part(commit_parts, part_idx)
        else:
            target.value = source.value


def _value_for_part(parts: Sequence[Any], part_idx: int) -> Any:
    """返回某个拆分行应写入的单元格值。"""
    if part_idx < len(parts):
        return parts[part_idx]
    return None


def _should_log_progress(processed: int, total: int, interval: int) -> bool:
    """判断是否需要输出长任务进度日志。"""
    if interval <= 0 or total <= 0:
        return False
    return processed % interval == 0 or processed == total


def split_worksheet(
    worksheet: Any, progress_interval: int = DEFAULT_PROGRESS_INTERVAL
) -> Dict[str, int]:
    """拆分单个 worksheet 中 Commit信息 列包含换行的行。"""
    stats = {
        "rows_scanned": 0,
        "rows_split": 0,
        "rows_added": 0,
        "cells_split": 0,
    }
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    commit_column = _find_commit_info_column(worksheet, max_column)
    if commit_column is None:
        LOG.warning(
            "跳过工作表: sheet=%s, reason=未找到 %s 表头",
            worksheet.title,
            COMMIT_INFO_HEADER,
        )
        return stats

    data_rows = max(0, max_row - 1)
    split_rows: Dict[int, List[Any]] = {}
    started_at = monotonic()

    LOG.info(
        "开始扫描工作表: sheet=%s, data_rows=%d, columns=%d, commit_column=%d",
        worksheet.title,
        data_rows,
        max_column,
        commit_column,
    )

    for row_idx in range(2, max_row + 1):
        stats["rows_scanned"] += 1
        commit_parts = _split_cell_value(
            worksheet.cell(row=row_idx, column=commit_column).value
        )
        if len(commit_parts) > 1:
            rows_to_add = len(commit_parts) - 1
            split_rows[row_idx] = commit_parts
            stats["rows_split"] += 1
            stats["rows_added"] += rows_to_add
            stats["cells_split"] += 1

        if _should_log_progress(stats["rows_scanned"], data_rows, progress_interval):
            LOG.info(
                "扫描进度: sheet=%s, scanned=%d/%d, split_rows=%d, "
                "added_rows=%d, elapsed=%.1fs",
                worksheet.title,
                stats["rows_scanned"],
                data_rows,
                stats["rows_split"],
                stats["rows_added"],
                monotonic() - started_at,
            )

    LOG.info(
        "扫描完成: sheet=%s, scanned=%d, split_rows=%d, added_rows=%d, "
        "split_cells=%d, elapsed=%.1fs",
        worksheet.title,
        stats["rows_scanned"],
        stats["rows_split"],
        stats["rows_added"],
        stats["cells_split"],
        monotonic() - started_at,
    )

    if not split_rows:
        LOG.info("无需拆分工作表: sheet=%s", worksheet.title)
        return stats

    rewrite_started_at = monotonic()
    final_rows = max_row + stats["rows_added"]
    LOG.info(
        "开始重写工作表: sheet=%s, original_rows=%d, final_rows=%d, commit_column=%d",
        worksheet.title,
        max_row,
        final_rows,
        commit_column,
    )

    added_after = 0
    rows_rewritten = 0
    for row_idx in range(max_row, 1, -1):
        commit_parts = split_rows.get(row_idx)
        rows_to_add = 0 if commit_parts is None else len(commit_parts) - 1
        added_before = stats["rows_added"] - added_after - rows_to_add
        target_start = row_idx + added_before

        if commit_parts is None:
            if target_start != row_idx:
                _copy_row_values_and_format(
                    worksheet, row_idx, target_start, max_column
                )
        else:
            for offset in range(len(commit_parts)):
                _copy_row_values_and_format(
                    worksheet,
                    row_idx,
                    target_start + offset,
                    max_column,
                    commit_column,
                    commit_parts,
                    offset,
                )

        added_after += rows_to_add
        rows_rewritten += 1
        if _should_log_progress(rows_rewritten, data_rows, progress_interval):
            LOG.info(
                "重写进度: sheet=%s, source_rows=%d/%d, elapsed=%.1fs",
                worksheet.title,
                rows_rewritten,
                data_rows,
                monotonic() - rewrite_started_at,
            )

    LOG.info(
        "重写完成: sheet=%s, final_rows=%d, elapsed=%.1fs",
        worksheet.title,
        final_rows,
        monotonic() - rewrite_started_at,
    )

    return stats


def _merge_stats(total: Dict[str, int], current: Dict[str, int]) -> None:
    """合并统计信息。"""
    for key, value in current.items():
        total[key] = total.get(key, 0) + value


def split_workbook(
    input_file: Path,
    output_file: Path,
    progress_interval: int = DEFAULT_PROGRESS_INTERVAL,
) -> Dict[str, int]:
    """读取、拆分并保存工作簿。"""
    openpyxl = _require_openpyxl()
    keep_vba = input_file.suffix.lower() == ".xlsm"
    started_at = monotonic()
    LOG.info("开始读取工作簿: input=%s, keep_vba=%s", input_file, keep_vba)
    workbook = openpyxl.load_workbook(str(input_file), keep_vba=keep_vba)
    LOG.info(
        "读取工作簿完成: sheets=%d, elapsed=%.1fs",
        len(workbook.worksheets),
        monotonic() - started_at,
    )

    stats = {
        "sheets": 0,
        "rows_scanned": 0,
        "rows_split": 0,
        "rows_added": 0,
        "cells_split": 0,
    }
    for worksheet in workbook.worksheets:
        stats["sheets"] += 1
        _merge_stats(stats, split_worksheet(worksheet, progress_interval))

    save_started_at = monotonic()
    LOG.info("开始保存工作簿: output=%s", output_file)
    workbook.save(str(output_file))
    LOG.info(
        "保存工作簿完成: output=%s, elapsed=%.1fs",
        output_file,
        monotonic() - save_started_at,
    )
    LOG.info("全部处理完成: elapsed=%.1fs", monotonic() - started_at)
    return stats


def collect_commits(input_file: Path) -> Dict[str, Any]:
    """从 Commit信息 列收集 patch_tool.py 可读取的 commit 列表。"""
    openpyxl = _require_openpyxl()
    started_at = monotonic()
    LOG.info("开始读取工作簿: input=%s, mode=export-commits", input_file)
    workbook = openpyxl.load_workbook(str(input_file), read_only=True, data_only=False)
    LOG.info(
        "读取工作簿完成: sheets=%d, elapsed=%.1fs",
        len(workbook.worksheets),
        monotonic() - started_at,
    )

    stats = {
        "sheets": 0,
        "sheets_skipped": 0,
        "rows_scanned": 0,
        "commits_exported": 0,
        "duplicates_skipped": 0,
    }
    entries: List[str] = []
    seen: Set[str] = set()

    try:
        for worksheet in workbook.worksheets:
            stats["sheets"] += 1
            commit_column = _find_commit_info_column(
                worksheet, worksheet.max_column
            )
            if commit_column is None:
                stats["sheets_skipped"] += 1
                LOG.warning(
                    "跳过工作表: sheet=%s, reason=未找到 %s 表头",
                    worksheet.title,
                    COMMIT_INFO_HEADER,
                )
                continue

            sheet_started_at = monotonic()
            sheet_exported = 0
            sheet_duplicates = 0
            LOG.info(
                "开始导出工作表: sheet=%s, rows=%d, commit_column=%d",
                worksheet.title,
                max(0, worksheet.max_row - 1),
                commit_column,
            )

            for row_values in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                max_col=commit_column,
                values_only=True,
            ):
                stats["rows_scanned"] += 1
                if len(row_values) < commit_column:
                    continue
                for part in _split_cell_value(row_values[commit_column - 1]):
                    commit = str(part).strip()
                    if not commit:
                        continue
                    if commit in seen:
                        stats["duplicates_skipped"] += 1
                        sheet_duplicates += 1
                        continue
                    seen.add(commit)
                    entries.append(commit)
                    stats["commits_exported"] += 1
                    sheet_exported += 1

            LOG.info(
                "导出工作表完成: sheet=%s, exported=%d, duplicates=%d, "
                "elapsed=%.1fs",
                worksheet.title,
                sheet_exported,
                sheet_duplicates,
                monotonic() - sheet_started_at,
            )
    finally:
        workbook.close()

    LOG.info(
        "全部收集完成: commits=%d, elapsed=%.1fs",
        len(entries),
        monotonic() - started_at,
    )
    return {"commits": entries, "stats": stats}


def export_commits(input_file: Path, output_file: Path) -> Dict[str, int]:
    """从 Commit信息 列导出 patch_tool.py 可读取的文本列表。"""
    started_at = monotonic()
    result = collect_commits(input_file)
    entries = result["commits"]
    stats = result["stats"]

    output_started_at = monotonic()
    LOG.info("开始写出 commit 列表: output=%s", output_file)
    with output_file.open("w", encoding="utf-8", newline="\n") as f:
        for entry in entries:
            f.write("{0}\n".format(entry))

    LOG.info(
        "写出 commit 列表完成: output=%s, commits=%d, elapsed=%.1fs",
        output_file,
        stats["commits_exported"],
        monotonic() - output_started_at,
    )
    LOG.info("全部导出完成: elapsed=%.1fs", monotonic() - started_at)
    return stats


def _validate_input_excel_path(input_file: Path) -> Optional[str]:
    """校验输入 Excel 路径。"""
    if not input_file.exists():
        return "输入文件不存在: {0}".format(input_file)
    if not input_file.is_file():
        return "输入路径不是普通文件: {0}".format(input_file)
    if input_file.suffix.lower() not in SUPPORTED_SUFFIXES:
        return "不支持的输入格式，仅支持 .xlsx/.xlsm: {0}".format(input_file)
    return None


def _validate_paths(input_file: Path, output_file: Path) -> Optional[str]:
    """校验输入输出路径。"""
    input_error = _validate_input_excel_path(input_file)
    if input_error:
        return input_error
    if output_file.suffix.lower() not in SUPPORTED_SUFFIXES:
        return "不支持的输出格式，仅支持 .xlsx/.xlsm: {0}".format(output_file)
    if output_file.parent and not output_file.parent.exists():
        return "输出目录不存在: {0}".format(output_file.parent)
    return None


def _validate_export_paths(input_file: Path, output_file: Path) -> Optional[str]:
    """校验 export-commits 输入输出路径。"""
    input_error = _validate_input_excel_path(input_file)
    if input_error:
        return input_error
    if output_file.parent and not output_file.parent.exists():
        return "输出目录不存在: {0}".format(output_file.parent)
    return None


def build_parser(prog: Optional[str] = None) -> argparse.ArgumentParser:
    """构造默认拆分命令参数解析器。"""
    parser = argparse.ArgumentParser(
        prog=prog,
        description="按 Commit信息 列换行符拆分 Excel 行",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("input", metavar="INPUT", help="输入 Excel 文件（.xlsx/.xlsm）")
    parser.add_argument("output", metavar="OUTPUT", help="输出 Excel 文件（.xlsx/.xlsm）")
    parser.add_argument(
        "--log-interval",
        type=int,
        default=DEFAULT_PROGRESS_INTERVAL,
        metavar="ROWS",
        help=(
            "扫描/重写进度日志的行间隔，0 表示关闭进度日志 "
            "（默认 {0}）".format(DEFAULT_PROGRESS_INTERVAL)
        ),
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="输出 debug 日志",
    )
    return parser


def build_export_commits_parser(prog: Optional[str] = None) -> argparse.ArgumentParser:
    """构造 export-commits 子命令参数解析器。"""
    parser = argparse.ArgumentParser(
        prog=prog,
        description="从 Excel Commit信息 列导出 patch_tool.py 可读取的 commit 列表",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("input", metavar="INPUT", help="输入 Excel 文件（.xlsx/.xlsm）")
    parser.add_argument("output", metavar="OUTPUT", help="输出文本文件")
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="输出 debug 日志",
    )
    return parser


def _configure_logging(verbose: bool) -> None:
    """初始化日志。"""
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(message)s",
    )


def cmd_split(args: argparse.Namespace) -> int:
    """执行默认 Excel 拆分命令。"""
    input_file = Path(args.input)
    output_file = Path(args.output)

    error = _validate_paths(input_file, output_file)
    if error:
        print("错误: {0}".format(error), file=sys.stderr)
        return 1

    try:
        stats = split_workbook(input_file, output_file, args.log_interval)
    except RuntimeError as exc:
        print("错误: {0}".format(exc), file=sys.stderr)
        return 1
    except OSError as exc:
        print("错误: 无法处理 Excel 文件: {0}".format(exc), file=sys.stderr)
        return 1

    print(
        "处理完成: sheets={sheets}, scanned_rows={rows_scanned}, "
        "split_rows={rows_split}, added_rows={rows_added}, split_cells={cells_split}".format(
            **stats
        )
    )
    print("输出文件: {0}".format(output_file))
    return 0


def cmd_export_commits(args: argparse.Namespace) -> int:
    """执行 export-commits 子命令。"""
    input_file = Path(args.input)
    output_file = Path(args.output)

    error = _validate_export_paths(input_file, output_file)
    if error:
        print("错误: {0}".format(error), file=sys.stderr)
        return 1

    try:
        stats = export_commits(input_file, output_file)
    except RuntimeError as exc:
        print("错误: {0}".format(exc), file=sys.stderr)
        return 1
    except OSError as exc:
        print("错误: 无法导出 commit 列表: {0}".format(exc), file=sys.stderr)
        return 1

    print(
        "导出完成: sheets={sheets}, skipped_sheets={sheets_skipped}, "
        "scanned_rows={rows_scanned}, commits={commits_exported}, "
        "duplicates_skipped={duplicates_skipped}".format(**stats)
    )
    print("输出文件: {0}".format(output_file))
    return 0


def main(argv: Optional[Sequence[str]] = None) -> int:
    """程序入口。"""
    raw_args = list(sys.argv[1:] if argv is None else argv)
    if raw_args and raw_args[0] == "export-commits":
        parser = build_export_commits_parser(
            prog="{0} export-commits".format(Path(sys.argv[0]).name)
        )
        args = parser.parse_args(raw_args[1:])
        _configure_logging(args.verbose)
        return cmd_export_commits(args)

    args = build_parser(prog=Path(sys.argv[0]).name).parse_args(raw_args)
    _configure_logging(args.verbose)
    return cmd_split(args)


if __name__ == "__main__":
    sys.exit(main())
